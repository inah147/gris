import io
import os
import re
import unicodedata
from datetime import datetime, timedelta

import frappe
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


def slugify(text: str) -> str:
	"""Remove acentos e caracteres especiais, retorna texto em lowercase com underscores."""
	if not text:
		return ""
	text = unicodedata.normalize("NFKD", text)
	text = text.encode("ascii", "ignore").decode("ascii")
	text = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_")
	return text.lower()


def get_instituicoes():
	"""Retorna lista de nomes de instituições financeiras."""
	instituicoes = frappe.get_all("Instituicao Financeira", fields=["name", "nome"], order_by="nome asc")
	return [i.get("nome") or i.get("name") for i in instituicoes]


def get_transacoes_periodo(data_inicio=None, data_fim=None):
	"""Busca transações do período filtrado, excluindo método Dinheiro."""
	filters = [["Transacao Extrato Geral", "metodo", "!=", "Dinheiro"]]

	if data_inicio:
		filters.append(["Transacao Extrato Geral", "data_deposito", ">=", data_inicio])
	if data_fim:
		filters.append(["Transacao Extrato Geral", "data_deposito", "<=", data_fim])

	transacoes = frappe.get_all(
		"Transacao Extrato Geral",
		fields=[
			"data_deposito",
			"instituicao",
			"carteira",
			"centro_de_custo",
			"valor",
			"valor_absoluto",
			"debito_credito",
			"descricao",
			"categoria",
			"nome_atividade",
			"observacoes",
			"metodo",
		],
		filters=filters,
		order_by="data_deposito asc",
		limit_page_length=0,
	)

	# Converter para dict e filtrar dinheiro (segurança adicional)
	rows = [dict(t) for t in transacoes]
	rows = [r for r in rows if str(r.get("metodo") or "").strip().lower() != "dinheiro"]
	rows = [
		r
		for r in rows
		if not str(r.get("descricao") or "").strip().lower().startswith("pagamento em dinheiro")
	]

	return rows


def get_saldos_iniciais_carteiras(instituicoes_nomes):
	"""Busca saldo inicial de todas as carteiras agrupado por instituição."""
	saldos = {inst: 0.0 for inst in instituicoes_nomes}

	try:
		res = frappe.db.sql(
			"""
			SELECT instituicao_financeira,
			       SUM(COALESCE(saldo_inicial, 0)) AS saldo
			FROM `tabCarteira`
			GROUP BY instituicao_financeira
			""",
			as_dict=True,
		)
		for row in res:
			inst = row.get("instituicao_financeira")
			if inst in saldos:
				saldos[inst] = float(row.get("saldo") or 0.0)
	except Exception:
		pass

	return saldos


def get_transacoes_anteriores(data_inicio, instituicoes_nomes):
	"""Busca e soma transações anteriores ao período (para cálculo de saldo inicial)."""
	saldos = {inst: 0.0 for inst in instituicoes_nomes}

	if not data_inicio:
		return saldos

	try:
		filters_anteriores = [
			["Transacao Extrato Geral", "data_deposito", "<", data_inicio],
			["Transacao Extrato Geral", "metodo", "!=", "Dinheiro"],
		]
		transacoes_anteriores = frappe.get_all(
			"Transacao Extrato Geral",
			fields=["instituicao", "valor"],
			filters=filters_anteriores,
			limit_page_length=0,
		)
		for t in transacoes_anteriores:
			inst = t.get("instituicao")
			if inst in saldos:
				valor = t.get("valor") or 0.0
				saldos[inst] += valor
	except Exception:
		pass

	return saldos


def calcular_saldos_iniciais(instituicoes_nomes, data_inicio=None):
	"""Calcula saldo inicial = saldo_inicial da carteira + transações anteriores ao período."""
	saldos_carteiras = get_saldos_iniciais_carteiras(instituicoes_nomes)
	saldos_anteriores = get_transacoes_anteriores(data_inicio, instituicoes_nomes)

	# Combinar ambos
	saldos_iniciais = {inst: 0.0 for inst in instituicoes_nomes}
	for inst in instituicoes_nomes:
		saldos_iniciais[inst] = saldos_carteiras.get(inst, 0.0) + saldos_anteriores.get(inst, 0.0)

	return saldos_iniciais


def calcular_saldos_por_linha(rows, instituicoes_nomes, saldos_iniciais):
	"""Calcula o saldo acumulado de cada instituição em cada linha de transação."""
	saldos = saldos_iniciais.copy()

	for r in rows:
		inst = r.get("instituicao")
		valor = r.get("valor") or 0.0

		# Atualizar saldo da instituição (valor já tem sinal correto)
		if inst in saldos:
			saldos[inst] += valor

		# Adicionar saldo de TODAS as instituições nesta linha
		for i in instituicoes_nomes:
			r[i] = saldos.get(i)

	return rows


def get_template_path():
	"""Localiza e retorna o caminho completo do arquivo de template."""
	template_name = "template_relatorio_contabil.xlsx"

	files = frappe.get_all(
		"File",
		filters={"file_name": template_name},
		fields=["name", "file_url", "is_private"],
		limit_page_length=1,
	)

	if not files:
		frappe.throw(f"Template '{template_name}' não encontrado como File. Faça upload com este nome.")

	file_doc = frappe.get_doc("File", files[0].get("name"))
	file_url = file_doc.file_url
	site_path = frappe.get_site_path()
	template_path = None

	# Resolver path local
	if file_url.startswith("/files/"):
		template_path = os.path.join(site_path, "public", file_url.lstrip("/"))
	elif file_url.startswith("/private/") or file_url.startswith("/private/files/"):
		template_path = os.path.join(site_path, file_url.lstrip("/"))
	else:
		try:
			template_path = file_doc.get_full_path()
		except Exception:
			pass

	if not template_path or not os.path.exists(template_path):
		frappe.throw(f"Arquivo de template não encontrado no filesystem: {template_path}")

	return template_path


def find_start_cell(wb):
	"""Localiza a célula inicial para preenchimento de dados (<<START_CELL>>)."""
	token_start = "<<START_CELL>>"

	# Procurar token literal em todas as planilhas
	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		max_r = ws.max_row or 100
		max_c = ws.max_column or 50

		for row in ws.iter_rows(min_row=1, max_row=max(1, max_r), min_col=1, max_col=max(1, max_c)):
			for cell in row:
				if isinstance(cell.value, str) and token_start in cell.value:
					# Remover token da célula
					new_val = cell.value.replace(token_start, "").strip()
					cell.value = new_val if new_val != "" else None
					return cell.row, cell.column, ws

	# Fallback: tentar nome definido START_CELL
	try:
		for defined in wb.defined_names.definedName:
			if defined.name == "START_CELL":
				dests = list(defined.destinations)
				if dests:
					sheet_name, coord = dests[0]
					if sheet_name in wb.sheetnames:
						col_letter, row_num = coordinate_from_string(coord)
						return int(row_num), column_index_from_string(col_letter), wb[sheet_name]
	except Exception:
		pass

	# Default: linha 2, coluna 1, planilha ativa
	return 2, 1, wb.active


def get_metadata_replacements(data_inicio, data_fim):
	"""Prepara dicionário com substituições de metadados."""
	now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

	try:
		export_user = frappe.get_value("User", frappe.session.user, "full_name") or frappe.session.user
	except Exception:
		export_user = frappe.session.user

	return {
		"<<start_date>>": data_inicio or "-",
		"<<end_date>>": data_fim or "-",
		"<<export_date>>": now_str or "-",
		"<<export_user>>": export_user or "-",
	}


def replace_tokens_in_workbook(wb, meta_replacements, instituicoes_nomes, saldos_iniciais):
	"""Substitui todos os tokens de metadados e saldos no workbook."""
	for sheet_name in wb.sheetnames:
		ws = wb[sheet_name]
		max_r = ws.max_row or 100
		max_c = ws.max_column or 50

		for row in ws.iter_rows(min_row=1, max_row=max(1, max_r), min_col=1, max_col=max(1, max_c)):
			for cell in row:
				if not isinstance(cell.value, str):
					continue

				v = cell.value
				replaced = False

				# Substituir tokens de metadados
				for token, repl in meta_replacements.items():
					if token in v:
						v = v.replace(token, str(repl))
						replaced = True

				# Substituir tokens de saldo por instituição
				for inst in instituicoes_nomes:
					slug = slugify(inst)
					token = f"<<saldo_{slug}>>"

					if token in v:
						saldo_val = saldos_iniciais.get(inst)

						if v.strip() == token:
							# Célula contém apenas o token: escrever valor numérico
							if saldo_val is None:
								cell.value = "-"
							else:
								try:
									cell.value = float(saldo_val)
									cell.number_format = "#,##0.00"
								except Exception:
									cell.value = saldo_val
							replaced = True
							break
						else:
							# Token embebido em texto
							if saldo_val is None:
								v = v.replace(token, "-")
							else:
								v = v.replace(token, f"{saldo_val:,.2f}")
							replaced = True

				if replaced and isinstance(cell.value, str):
					cell.value = v


def format_cell_value(value, default="-"):
	"""Retorna valor ou default se vazio/None."""
	return value if value not in (None, "") else default


def write_transaction_row(ws, row_idx, base_col, transaction, dyn_base_col, instituicoes_nomes, last_saldos):
	"""Escreve uma linha de transação no worksheet."""
	# Data
	dt = transaction.get("data_deposito")
	if isinstance(dt, str):
		try:
			dt = datetime.fromisoformat(dt)
		except Exception:
			pass
	ws.cell(row=row_idx, column=base_col, value=dt)

	# Campos textuais
	ws.cell(row=row_idx, column=base_col + 1, value=format_cell_value(transaction.get("instituicao")))
	ws.cell(row=row_idx, column=base_col + 2, value=format_cell_value(transaction.get("carteira")))
	ws.cell(row=row_idx, column=base_col + 3, value=format_cell_value(transaction.get("centro_de_custo")))

	# Valor (usar valor_absoluto para exibição)
	val_cell = ws.cell(row=row_idx, column=base_col + 4, value=transaction.get("valor_absoluto") or 0)
	val_cell.number_format = "#,##0.00"

	# Demais campos
	ws.cell(row=row_idx, column=base_col + 5, value=format_cell_value(transaction.get("debito_credito")))
	ws.cell(row=row_idx, column=base_col + 6, value=format_cell_value(transaction.get("descricao")))
	ws.cell(row=row_idx, column=base_col + 7, value=format_cell_value(transaction.get("categoria")))
	ws.cell(row=row_idx, column=base_col + 8, value=format_cell_value(transaction.get("nome_atividade")))
	ws.cell(row=row_idx, column=base_col + 9, value=format_cell_value(transaction.get("observacoes")))

	# Colunas dinâmicas de saldo por instituição
	for j, inst in enumerate(instituicoes_nomes):
		col_idx = dyn_base_col + j
		saldo = transaction.get(inst)

		if saldo in (None, ""):
			# Sem mudança nesta instituição: carregar último saldo conhecido
			saldo = last_saldos.get(inst)

		if saldo is None:
			ws.cell(row=row_idx, column=col_idx, value="-")
		else:
			try:
				saldo = float(saldo)
				cell = ws.cell(row=row_idx, column=col_idx, value=saldo)
				cell.number_format = "#,##0.00"
				last_saldos[inst] = saldo  # Atualizar último saldo conhecido
			except Exception:
				ws.cell(row=row_idx, column=col_idx, value=saldo)


def apply_alternating_row_colors(ws, row_idx, start_row, base_col, dyn_base_col, num_institutions):
	"""Aplica cor alternada em linhas (cinza claro em linhas pares)."""
	offset = row_idx - start_row
	if offset % 2 == 0:
		light_fill = PatternFill(start_color="FFF7F7F7", end_color="FFF7F7F7", fill_type="solid")
		for c in range(base_col, dyn_base_col + num_institutions):
			ws.cell(row=row_idx, column=c).fill = light_fill


def write_data_to_workbook(wb, start_row, start_col, start_ws, rows, instituicoes_nomes, saldos_iniciais):
	"""Escreve todas as linhas de transação no workbook."""
	data_base_col = start_col
	dyn_base_col = data_base_col + 10

	# Preparar últimos saldos conhecidos
	last_saldos = {
		inst: (saldos_iniciais.get(inst) if saldos_iniciais.get(inst) is not None else 0.0)
		for inst in instituicoes_nomes
	}

	for ri, transaction in enumerate(rows, start=start_row):
		write_transaction_row(
			start_ws, ri, data_base_col, transaction, dyn_base_col, instituicoes_nomes, last_saldos
		)
		apply_alternating_row_colors(
			start_ws, ri, start_row, data_base_col, dyn_base_col, len(instituicoes_nomes)
		)


def create_excel_response(wb):
	"""Salva workbook em bytes e prepara resposta de download."""
	out = io.BytesIO()
	wb.save(out)
	out.seek(0)

	filename = f"relatorio_contabil_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = out.read()
	frappe.local.response.type = "download"


@frappe.whitelist()
def export_relatorio_contabil(data_inicio=None, data_fim=None):
	"""Exporta o relatório contábil preenchendo o template Excel.

	Args:
		data_inicio: Data inicial do período (opcional)
		data_fim: Data final do período (opcional)

	Returns:
		Arquivo Excel para download
	"""
	if frappe.session.user == "Guest":
		frappe.throw("Login necessário", frappe.PermissionError)

	# 1. Buscar dados
	instituicoes_nomes = get_instituicoes()
	rows = get_transacoes_periodo(data_inicio, data_fim)

	# 2. Calcular saldos
	saldos_iniciais = calcular_saldos_iniciais(instituicoes_nomes, data_inicio)
	rows = calcular_saldos_por_linha(rows, instituicoes_nomes, saldos_iniciais)

	# 3. Carregar template Excel
	template_path = get_template_path()
	wb = load_workbook(template_path)

	# 4. Localizar célula inicial
	start_row, start_col, start_ws = find_start_cell(wb)

	# 5. Substituir tokens de metadados e saldos
	meta_replacements = get_metadata_replacements(data_inicio, data_fim)
	replace_tokens_in_workbook(wb, meta_replacements, instituicoes_nomes, saldos_iniciais)

	# 6. Escrever dados das transações
	write_data_to_workbook(wb, start_row, start_col, start_ws, rows, instituicoes_nomes, saldos_iniciais)

	# 7. Gerar resposta de download
	create_excel_response(wb)
